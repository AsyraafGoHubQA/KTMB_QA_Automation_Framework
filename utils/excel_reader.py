from openpyxl import load_workbook


def read_login_data(file_path, tc=None):

    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        print("ROW =", row)

        record = {
            "tc": row[0],
            "username": row[1],
            "password": row[2],
            "expected": row[3]
        }

        print("RECORD =", record)

        if tc is not None:

            if str(record["tc"]).strip() == str(tc).strip():
                data.append(record)

        else:
            data.append(record)

    return data