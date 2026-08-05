from openpyxl import load_workbook


def read_login_data(file_path, tc=None):

    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []

    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):

        record = dict(zip(headers, row))

        # Return only the selected TC
        if tc:

            if record["tc"] == tc:
                return [record]

        else:
            data.append(record)

    return data